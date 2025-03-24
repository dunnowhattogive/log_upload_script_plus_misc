#!/usr/bin/env python3
import os
import json
import pandas as pd
import socket
import uuid
from smbprotocol.connection import Connection
from smbprotocol.session import Session
from smbprotocol.tree import TreeConnect
from smbprotocol.open import Open, CreateDisposition, FileAttributes, ImpersonationLevel, ShareAccess, CreateOptions
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image
import logging
import logging.handlers
if os.name != 'nt':
    import syslog
import concurrent.futures
import argparse

# Configure base directory and server details
BASE_DIR = os.path.expanduser("~/Documents/")
SERVER_URL = "172.22.41.14/AICPROD-FILE01/DATA"  # Replace with your server URL
SERVER_IP = "172.21.41.14"
HEADERS = {"Authorization": "SVC_FileExport@aicprod.local Ju$5LKEbreU!"}  # Replace with your API token or authentication header

# SMB server details
# SERVER_IP = "192.168.1.10"
SERVER_IP = "172.22.41.14"
# SHARE_NAME = "shared_folder"
SHARE_NAME = "AIC-Helsinge"
# DESTINATION_PATH = "uploads"
DESTINATION_PATH = "SMART_CATCH_MINI"
# USERNAME = "your_username"
# USERNAME = "SVC_FileExport@aicprod.local"
USERNAME = "aicprod\AIC-SVC_FileExport"
# PASSWORD = "your_password"
PASSWORD = "Ju$5LKEbreU!"

# SMB server details
SERVER_NAME = "AICPROD-FILE01"
# SERVER_NAME = socket.gethostname()

def setup_logging():
    """Configure logging to syslog and console, and provide a function to log messages."""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    if os.name == 'nt':  # Windows
        # Add a handler to log to the Windows Event Log
        event_log_handler = logging.handlers.NTEventLogHandler("Python Application")
        event_log_formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s', datefmt='%b %d %H:%M:%S')
        event_log_handler.setFormatter(event_log_formatter)
        logger.addHandler(event_log_handler)
    else:
        # Add a handler to log to /var/log/syslog
        syslog_handler = logging.handlers.SysLogHandler(address='/dev/log')
        syslog_formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s', datefmt='%b %d %H:%M:%S')
        syslog_handler.setFormatter(syslog_formatter)
        logger.addHandler(syslog_handler)

    # Also add a StreamHandler to log to console for debugging purposes
    console_handler = logging.StreamHandler()
    console_formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s', datefmt='%b %d %H:%M:%S')
    console_handler.setFormatter(console_formatter)
    logger.addHandler(console_handler)

    def log_message(level, message):
        """Log a message to both syslog and console."""
        if level == "info":
            logger.info(message)
            if os.name != 'nt':
                syslog.syslog(syslog.LOG_INFO, message)
        elif level == "error":
            logger.error(message)
            if os.name != 'nt':
                syslog.syslog(syslog.LOG_ERR, message)
        print(message)

    return log_message

log_message = setup_logging()

def parse_old_json_structure(test_data):
    """Parse the old JSON structure into a flat dictionary."""
    parsed_data = {
        "Device Type": test_data.get("devicetype"),
        "Firmware Version": test_data.get("firmwareversion"),
        "Hardware Info": test_data.get("hardwareinfo"),
    }
    
    for test, test_info in test_data.items():
        if isinstance(test_info, dict):
            if test == "testflow":
                # Check if all components are "true"
                all_pass = all(str(value).lower() == "true" for value in test_info.values())
                parsed_data[f"{test} Results"] = "PASS" if all_pass else "FAIL"
                if not all_pass:
                    # Collect the keys with "false" values
                    failures = [key for key, value in test_info.items() if str(value).lower() == "false"]
                    parsed_data[f"{test} Failures"] = ", ".join(failures)
                else:
                    parsed_data[f"{test} Failures"] = "N/A"
            else:
                if isinstance(test_info.get("results"), list):
                    parsed_data[f"{test} Results"] = ", ".join(test_info["results"])
                else:
                    parsed_data[f"{test} Results"] = test_info.get("results")
                parsed_data[f"{test} Failures"] = test_info.get("number_of_fails", "N/A")
                parsed_data[f"{test} Timestamp Start"] = test_info.get("timestamp_begin")
                parsed_data[f"{test} Timestamp End"] = test_info.get("timestamp_end")
    
    return parsed_data

def parse_new_json_structure(test_data):
    """Parse the new JSON structure into a flat dictionary."""
    parsed_data = {
        "All Data Dev": test_data.get("all_data_dev"),
        "All Data Mean": test_data.get("all_data_mean"),
        "Current Result": test_data.get("current_result"),
        "Current UA": test_data.get("current_ua"),
        "Current UA Std": test_data.get("current_ua_std"),
        "Location": test_data.get("location"),
        "Raw Data": test_data.get("raw_data"),
        "Test Result": test_data.get("test_result"),
        "Tester": test_data.get("tester"),
        "Timestamp Begin": datetime.fromtimestamp(test_data.get("timestamp_begin")).strftime('%Y-%m-%d %H:%M:%S'),
        "Timestamp End": datetime.fromtimestamp(test_data.get("timestamp_end")).strftime('%Y-%m-%d %H:%M:%S'),
        "Voltage MV": test_data.get("voltage_mv"),
        "Voltage Result": test_data.get("voltage_result")
    }
    
    # Determine the final result based on current_result and test_result
    if test_data.get("current_result") == "PASSED" and test_data.get("test_result") == "PASSED":
        parsed_data["Final Result"] = "PASS"
    else:
        parsed_data["Final Result"] = "FAIL"
    
    return parsed_data

def upload_file_smb(file_path, server_name, share_name, destination_path, username, password, port=445):
    """
    Uploads a file to an SMB share using smbprotocol.

    Parameters:
    - file_path: Local path of the file to be uploaded.
    - server_name: Hostname or NetBIOS name of the SMB server.
    - share_name: Name of the shared folder on the SMB server.
    - destination_path: Path within the share to upload the file.
    - username: Username for authentication.
    - password: Password for authentication.
    - port: Port for SMB (default is 445).
    """
    # Check if today is Friday
    if datetime.today().weekday() != 6:  # 6 corresponds to Sunday
        message = "File upload is only allowed on Sundays."
        log_message("info", message)
        return

    try:
        # Step 1: Establish an SMB connection with a unique GUID
        message = f"Connecting to SMB server {server_name} on port {port}..."
        log_message("info", message)
        conn = Connection(server_name=server_name, port=port, guid=uuid.uuid4())
        conn.connect()

        # Step 2: Start a session
        session = Session(connection=conn, username=username, password=password)
        session.connect()

        # Step 3: Connect to the SMB share
        tree = TreeConnect(session=session, share_name=f"\\\\{server_name}\\{share_name}")
        tree.connect()

        # Step 4: Resolve the remote file path
        destination_file_path = f"{destination_path}\\{file_path.split('/')[-1]}"

        # Step 5: Open a file handle for uploading
        file_handle = Open(tree, destination_file_path)
        file_handle.create(
            create_disposition=CreateDisposition.FILE_OVERWRITE_IF,
            file_attributes=FileAttributes.FILE_ATTRIBUTE_NORMAL,
            desired_access=0x001201BF,  # File write access
            impersonation_level=ImpersonationLevel.Impersonation,
            share_access=ShareAccess.FILE_SHARE_WRITE | ShareAccess.FILE_SHARE_READ,
            create_options=CreateOptions.FILE_NON_DIRECTORY_FILE
        )

        # Step 6: Read and write the file in chunks
        message = f"Uploading {file_path} to {destination_file_path}..."
        log_message("info", message)
        with open(file_path, "rb") as file:
            offset = 0
            while chunk := file.read(1024):  # Read in 1KB chunks
                file_handle.write(chunk, offset=offset)
                offset += len(chunk)

        # Step 7: Close the file handle
        file_handle.close()
        message = f"File {file_path} uploaded successfully to {destination_file_path}"
        log_message("info", message)

        # Step 8: Disconnect resources
        tree.disconnect()
        session.disconnect()
        conn.disconnect()

    except Exception as e:
        message = f"Error uploading {file_path} via SMB: {e}"
        log_message("error", message)

def process_json_file(json_path, new_json):
    """Process a single JSON file and return the record."""
    try:
        with open(json_path, 'r') as json_file:
            data = json.load(json_file)
        records = []
        for test_id in range(data["number_of_tests"]):
            if str(test_id) in data:
                if new_json:
                    record = parse_new_json_structure(data[str(test_id)])
                else:
                    record = parse_old_json_structure(data[str(test_id)])
                file_date = datetime.fromtimestamp(os.path.getmtime(json_path)).strftime('%Y-%m-%d %H:%M:%S')
                record = {"Date": file_date, "File Name": os.path.basename(json_path), **record}  # Add date and file name as the first columns
                records.append(record)
        return records
    except Exception as e:
        log_message("error", f"Failed to process {json_path}: {e}")
        return []

def get_default_directory(new_json):
    """Get the default directory based on the operating system and JSON structure."""
    if new_json:
        base_dir = os.path.join(BASE_DIR, "current_measurement_test_" + socket.gethostname())
    else:
        base_dir = os.path.join(BASE_DIR, "node_distributed_test_" + socket.gethostname())
    
    if os.name == 'nt':  # Windows
        base_dir = base_dir.replace('/', '\\')
        base_dir = os.path.normpath(base_dir)
    return base_dir

def order_sheets_by_week(writer):
    """Order the sheets in the workbook in ascending order of week number."""
    wb = writer.book
    sheet_names = wb.sheetnames
    week_sheets = [name for name in sheet_names if name.startswith("Week_")]
    week_sheets.sort(key=lambda x: int(x.split('-W')[1]))
    for sheet_name in week_sheets:
        sheet = wb[sheet_name]
        wb.remove(sheet)
        wb._add_sheet(sheet)

def process_directory(custom_dir=None, by_year=False, new_json=False):
    """Processes all JSON files in a folder and combines them into one Excel file per month or year."""
    log_message("info", "Starting process_directory function")
    if custom_dir:
        dir_path = custom_dir
    else:
        dir_path = get_default_directory(new_json)
    
    if not os.path.exists(dir_path):
        log_message("error", f"Directory {dir_path} does not exist.")
        return
    
    aggregated_records = {}  # Dictionary to hold data per month or week
    json_files = []

    for root, _, files in os.walk(dir_path):
        for file_name in files:
            if file_name.endswith(".json"):
                json_files.append(os.path.join(root, file_name))

    # Process JSON files in parallel
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_json_file, json_path, new_json) for json_path in json_files]
        for future in concurrent.futures.as_completed(futures):
            records = future.result()
            for record in records:
                record_date = datetime.strptime(record["Date"], '%Y-%m-%d %H:%M:%S')
                if by_year:
                    year, week, _ = record_date.isocalendar()
                    week = min(week, 52)  # Ensure week number does not exceed 52
                    time_key = f"{year}-W{week:02d}"
                else:
                    time_key = record_date.strftime('%Y-%m')
                if time_key not in aggregated_records:
                    aggregated_records[time_key] = []
                aggregated_records[time_key].append(record)

    # Save aggregated data to separate Excel files for each month or year
    for time_key, records in aggregated_records.items():
        if by_year:
            year, week = time_key.split('-W')
            week = min(int(week), 52)  # Ensure week number does not exceed 52
            aggregated_file_path = os.path.join(os.path.dirname(__file__), f"aggregated_results_{year}.xlsx")
            sheet_name = f"Week_{year}-W{week:02d}"
        else:
            aggregated_file_path = os.path.join(os.path.dirname(__file__), f"aggregated_results_{time_key}.xlsx")
            sheet_name = f"Month_{time_key}"
        
        try:
            # Check if the file exists and create it if it does not
            if not os.path.exists(aggregated_file_path):
                with pd.ExcelWriter(aggregated_file_path, engine='openpyxl') as writer:
                    df = pd.DataFrame(records)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    analyze_and_plot_results(writer, df, sheet_name, new_json)
                    if by_year:
                        order_sheets_by_week(writer)
            else:
                with pd.ExcelWriter(aggregated_file_path, engine='openpyxl', mode='a') as writer:
                    if sheet_name in writer.book.sheetnames:
                        log_message("error", f"Sheet '{sheet_name}' already exists in {aggregated_file_path}. Skipping.")
                        continue
                    df = pd.DataFrame(records)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    analyze_and_plot_results(writer, df, sheet_name, new_json)
                    if by_year:
                        order_sheets_by_week(writer)
            
            message = f"Aggregated Excel file created: {aggregated_file_path}"
            log_message("info", message)

            # Upload file via SMB
            upload_file_smb(aggregated_file_path, SERVER_NAME, SHARE_NAME, DESTINATION_PATH, USERNAME, PASSWORD)

        except Exception as e:
            log_message("error", f"Failed to create or upload the aggregated Excel file for {time_key}: {e}")

def analyze_and_plot_results(writer, df, sheet_name, new_json=False):
    """
    Analyzes the DataFrame for test results and plots a pie chart for pass and fail cases.
    Parameters:
    - writer: ExcelWriter object to write the plot.
    - df: DataFrame containing test results.
    - sheet_name: Name of the sheet to write the plot.
    - new_json: Boolean flag indicating if the new JSON structure is being used.
    """
    try:
        # Initialize counters for pass and fail cases
        pass_count = 0
        fail_count = 0
        indeterminate_count = 0
        failure_reasons = []
        failure_types_count = {}

        # Check each test result column for failures
        file_results = {}
        for _, row in df.iterrows():
            file_name = row["File Name"]
            if file_name not in file_results:
                file_results[file_name] = {"pass": 0, "fail": 0, "reasons": []}
            
            row_failures = 0
            reasons = []
            for column in df.columns:
                if "Failures" in column:
                    failures = row[column]
                    if pd.notna(failures) and failures != "N/A":
                        if str(failures).isdigit():
                            failure_count = int(failures)
                            if failure_count > 0:
                                row_failures += failure_count
                                reasons.append(f"{column}: {failures} failures")
                                failure_types_count[column] = failure_types_count.get(column, 0) + failure_count
                        else:
                            reasons.append(f"{column}: {failures}")
            # Check for failures in Current Result, Voltage Result, and both if new_json is True
            if new_json:
                current_failed = row.get("Current Result") == "FAILED"
                voltage_failed = row.get("Voltage Result") == "FAILED"
                if current_failed:
                    row_failures += 1
                    reasons.append("Current Result failed")
                    failure_types_count["Current Result"] = failure_types_count.get("Current Result", 0) + 1
                if voltage_failed:
                    row_failures += 1
                    reasons.append("Voltage Result failed")
                    failure_types_count["Voltage Result"] = failure_types_count.get("Voltage Result", 0) + 1
                if current_failed and voltage_failed:
                    failure_types_count["Both Current and Voltage Result"] = failure_types_count.get("Both Current and Voltage Result", 0) + 1
            else:
                if row.get("testflow Results") == "FAIL":
                    row_failures += 1
                    reasons.append("Testflow failed")
                    failure_types_count["Testflow"] = failure_types_count.get("Testflow", 0) + 1
                if row.get("testflow Failures") != "N/A":
                    reasons.append(f"Testflow Failures: {row.get('testflow Failures')}")
            
            if row_failures > 0:
                file_results[file_name]["fail"] += 1
                file_results[file_name]["reasons"].append(", ".join(reasons))
            else:
                file_results[file_name]["pass"] += 1
                file_results[file_name]["reasons"].append("None")

        # Determine the final result for each file
        for file_name, results in file_results.items():
            if results["pass"] > results["fail"]:
                pass_count += 1
            elif results["fail"] > results["pass"]:
                fail_count += 1
            else:
                indeterminate_count += 1

        # Add failure reasons to the DataFrame
        df["Failure Reasons"] = df.apply(lambda row: file_results[row["File Name"]]["reasons"].pop(0), axis=1)

        # Plot the pie chart
        import matplotlib.pyplot as plt

        labels = 'Pass', 'Fail', 'Indeterminate'
        sizes = [pass_count, fail_count, indeterminate_count]
        colors = ['#4CAF50', '#F44336', '#FFC107']
        explode = (0.1, 0, 0)  # explode the 1st slice (Pass)

        def autopct_format(pct, allvalues):
            absolute = int(round(pct/100.*sum(allvalues)))
            return f"{pct:.1f}% ({absolute})"

        plt.pie(sizes, explode=explode, labels=labels, colors=colors, autopct=lambda pct: autopct_format(pct, sizes),
                shadow=True, startangle=140)
        plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
        plt.title('Test Results: Pass vs Fail vs Indeterminate')

        # Add legend for failure types and make the text red
        legend_text = [f"{key}: {value} failures" for key, value in failure_types_count.items()]
        legend = plt.legend(legend_text, title="Failure Types", loc='lower left', fontsize='small', title_fontsize='medium', fancybox=True, shadow=True)
        plt.setp(legend.get_texts(), color='red')
        for handle in legend.legend_handles:
            handle.set_color('red')

        # Save the plot to a BytesIO object
        img_data = BytesIO()
        plt.savefig(img_data, format='png')
        plt.close()
        img_data.seek(0)

        # Load the Excel file and embed the image
        wb = writer.book
        ws = wb[sheet_name]
        img = Image(img_data)
        img.anchor = 'H2'  # Position the image at cell H2
        ws.add_image(img)

        # Color code the pass and fail cells
        pass_fill = openpyxl.styles.PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        fail_fill = openpyxl.styles.PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
        indeterminate_fill = openpyxl.styles.PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        duplicate_fill = openpyxl.styles.PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if "Failures" in ws.cell(row=1, column=cell.column).value:
                    if cell.value != "N/A" and str(cell.value).isdigit() and int(cell.value) > 0:
                        cell.fill = fail_fill
                    else:
                        cell.fill = pass_fill
                elif new_json and ws.cell(row=1, column=cell.column).value in ['Current Result', 'Voltage Result']:
                    if cell.value == "PASSED":
                        cell.fill = pass_fill
                    elif cell.value == "FAILED":
                        cell.fill = fail_fill
                elif not new_json and ws.cell(row=1, column=cell.column).value in ['testflow Results']:
                    if cell.value == "PASS":
                        cell.fill = pass_fill
                    elif cell.value == "FAIL":
                        cell.fill = fail_fill

        # Highlight duplicate file names in orange
        file_names = [ws.cell(row=row, column=2).value for row in range(2, ws.max_row + 1)]
        duplicates = set([name for name in file_names if file_names.count(name) > 1])
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=2).value in duplicates:
                ws.cell(row=row, column=2).fill = duplicate_fill

        # Add the "Failure Reasons" column to the worksheet
        ws.insert_cols(ws.max_column + 1)
        ws.cell(row=1, column=ws.max_column).value = "Failure Reasons"
        for idx, reason in enumerate(df["Failure Reasons"], start=2):
            ws.cell(row=idx, column=ws.max_column).value = reason

        # Color code the failure reasons column
        failure_reasons_col = ws.iter_cols(min_col=ws.max_column, max_col=ws.max_column, min_row=2, max_row=ws.max_row)
        for col in failure_reasons_col:
            for cell in col:
                if cell.value == "Indeterminate":
                    cell.fill = indeterminate_fill
                elif cell.value == "None":
                    cell.fill = pass_fill
                else:
                    cell.fill = fail_fill

        log_message("info", f"Pie chart embedded into Excel sheet: {sheet_name}")

    except Exception as e:
        log_message("error", f"Error analyzing or plotting results: {e}")

def main():
    parser = argparse.ArgumentParser(description="Process and upload JSON files to SMB server.")
    parser.add_argument("directory", nargs="?", default=BASE_DIR, help="Directory containing JSON files")
    parser.add_argument("--yearly", action="store_true", help="Aggregate data by year")
    parser.add_argument("--battery-test", action="store_true", help="Process JSON files with new structure")
    args = parser.parse_args()

    directory = args.directory
    yearly = args.yearly
    battery_test = args.battery_test

    log_message("info", "Starting the script")
    process_directory(directory, yearly, battery_test)
    log_message("info", "Script completed successfully")

if __name__ == "__main__":
    main()